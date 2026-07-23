"""
Excel Reporter — Generates 7-sheet Automation_Test_Report.xlsx from execution-results.json.
"""

import json
import os
import sys
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

STATUS_COLORS = {
    'PASSED':  'FF1DB954',  # Green
    'FAILED':  'FFDC3545',  # Red
    'SKIPPED': 'FFFFC107',  # Amber
    'BLOCKED': 'FF6C757D',  # Grey
}
HEADER_FILL   = PatternFill('solid', fgColor='FF1A1A2E')
HEADER_FONT   = Font(bold=True, color='FFFFFFFF', size=11)
TITLE_FONT    = Font(bold=True, color='FF1A1A2E', size=14)
BORDER_THIN   = Border(
    left=Side(style='thin', color='FFD0D0D0'),
    right=Side(style='thin', color='FFD0D0D0'),
    top=Side(style='thin', color='FFD0D0D0'),
    bottom=Side(style='thin', color='FFD0D0D0')
)


def _cell_style(ws, row, col, value, fill=None, font=None, align='left', border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if border:
        cell.border = BORDER_THIN
    return cell


def _write_test_sheet(ws, title: str, tests: list, cols: list):
    ws.title = title
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    _cell_style(ws, 1, 1, title, font=TITLE_FONT, align='center')
    ws.row_dimensions[1].height = 30

    # Header row
    for c_idx, col in enumerate(cols, 1):
        _cell_style(ws, 2, c_idx, col, fill=HEADER_FILL, font=HEADER_FONT, align='center')
    ws.row_dimensions[2].height = 20

    # Data rows
    for r_idx, test in enumerate(tests, 3):
        status = test.get('status', 'UNKNOWN')
        row_fill_color = STATUS_COLORS.get(status.upper(), 'FFFFFFFF')
        row_fill = PatternFill('solid', fgColor=row_fill_color + '33')  # 20% opacity

        row_data = [
            test.get('id', ''),
            test.get('module', ''),
            test.get('name', ''),
            test.get('priority', 'P2'),
            status,
            f"{test.get('duration_ms', 0):.1f}ms",
            test.get('failure_reason', ''),
            test.get('screenshot', ''),
        ]
        for c_idx, val in enumerate(row_data[:len(cols)], 1):
            cell = _cell_style(ws, r_idx, c_idx, val)
            cell.fill = row_fill
            if col == 'Status':
                status_fill = PatternFill('solid', fgColor=row_fill_color)
                cell.fill = status_fill
                cell.font = Font(bold=True, color='FFFFFFFF')

    # Column widths
    col_widths = [12, 20, 45, 8, 10, 12, 40, 35]
    for i, w in enumerate(col_widths[:len(cols)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_excel_report(results: dict, output_dir: str):
    os.makedirs(output_dir, exist_ok=True)
    metrics = results.get('metrics', {})
    tests   = results.get('test_cases', [])

    passed  = [t for t in tests if t.get('status', '').upper() == 'PASSED']
    failed  = [t for t in tests if t.get('status', '').upper() == 'FAILED']
    skipped = [t for t in tests if t.get('status', '').upper() == 'SKIPPED']

    wb = Workbook()
    ALL_COLS    = ['Test ID', 'Module', 'Test Name', 'Priority', 'Status', 'Duration', 'Failure Reason', 'Screenshot']
    SIMPLE_COLS = ['Test ID', 'Module', 'Test Name', 'Priority', 'Status', 'Duration', 'Failure Reason', 'Screenshot']

    # Sheet 1: All Executed
    ws1 = wb.active
    _write_test_sheet(ws1, 'All Executed Tests', tests, ALL_COLS)

    # Sheet 2: Passed
    ws2 = wb.create_sheet('Passed Tests')
    _write_test_sheet(ws2, 'Passed Tests', passed, SIMPLE_COLS)

    # Sheet 3: Failed
    ws3 = wb.create_sheet('Failed Tests')
    _write_test_sheet(ws3, 'Failed Tests', failed, SIMPLE_COLS)

    # Sheet 4: Skipped
    ws4 = wb.create_sheet('Skipped Tests')
    _write_test_sheet(ws4, 'Skipped Tests', skipped, SIMPLE_COLS)

    # Sheet 5: Execution Metrics
    ws5 = wb.create_sheet('Execution Metrics')
    ws5.title = 'Execution Metrics'
    metric_data = [
        ('Total Test Cases', metrics.get('total', len(tests))),
        ('Passed', metrics.get('passed', len(passed))),
        ('Failed', metrics.get('failed', len(failed))),
        ('Skipped', metrics.get('skipped', len(skipped))),
        ('Pass Rate (%)', f"{metrics.get('pass_rate', 0):.1f}%"),
        ('Fail Rate (%)', f"{metrics.get('fail_rate', 0):.1f}%"),
        ('Total Duration', metrics.get('duration', 'N/A')),
        ('Execution Start', metrics.get('start_time', 'N/A')),
        ('Execution End', metrics.get('end_time', 'N/A')),
        ('App Version', metrics.get('app_version', '1.0.0')),
        ('Android Version', metrics.get('android_version', '14')),
        ('Device', metrics.get('device', 'E2E_Emulator')),
    ]
    _cell_style(ws5, 1, 1, 'Execution Metrics Summary', font=TITLE_FONT)
    _cell_style(ws5, 2, 1, 'Metric', fill=HEADER_FILL, font=HEADER_FONT)
    _cell_style(ws5, 2, 2, 'Value', fill=HEADER_FILL, font=HEADER_FONT)
    for i, (k, v) in enumerate(metric_data, 3):
        _cell_style(ws5, i, 1, k)
        _cell_style(ws5, i, 2, str(v))
    ws5.column_dimensions['A'].width = 30
    ws5.column_dimensions['B'].width = 30

    # Sheet 6: Defect Summary by Module
    ws6 = wb.create_sheet('Defect Summary')
    module_failures = {}
    for t in failed:
        mod = t.get('module', 'Unknown')
        module_failures[mod] = module_failures.get(mod, 0) + 1
    _cell_style(ws6, 1, 1, 'Defect Summary by Module', font=TITLE_FONT)
    _cell_style(ws6, 2, 1, 'Module', fill=HEADER_FILL, font=HEADER_FONT)
    _cell_style(ws6, 2, 2, 'Failure Count', fill=HEADER_FILL, font=HEADER_FONT)
    for i, (mod, cnt) in enumerate(sorted(module_failures.items(), key=lambda x: -x[1]), 3):
        _cell_style(ws6, i, 1, mod)
        _cell_style(ws6, i, 2, cnt)
    ws6.column_dimensions['A'].width = 35
    ws6.column_dimensions['B'].width = 20

    # Sheet 7: Pass Rate by Module
    ws7 = wb.create_sheet('Pass Rate Summary')
    module_stats = {}
    for t in tests:
        mod = t.get('module', 'Unknown')
        if mod not in module_stats:
            module_stats[mod] = {'pass': 0, 'total': 0}
        module_stats[mod]['total'] += 1
        if t.get('status', '').upper() == 'PASSED':
            module_stats[mod]['pass'] += 1
    _cell_style(ws7, 1, 1, 'Pass Rate by Module', font=TITLE_FONT)
    _cell_style(ws7, 2, 1, 'Module', fill=HEADER_FILL, font=HEADER_FONT)
    _cell_style(ws7, 2, 2, 'Total', fill=HEADER_FILL, font=HEADER_FONT)
    _cell_style(ws7, 2, 3, 'Passed', fill=HEADER_FILL, font=HEADER_FONT)
    _cell_style(ws7, 2, 4, 'Pass Rate', fill=HEADER_FILL, font=HEADER_FONT)
    for i, (mod, s) in enumerate(sorted(module_stats.items()), 3):
        rate = (s['pass'] / s['total'] * 100) if s['total'] > 0 else 0
        _cell_style(ws7, i, 1, mod)
        _cell_style(ws7, i, 2, s['total'])
        _cell_style(ws7, i, 3, s['pass'])
        _cell_style(ws7, i, 4, f"{rate:.1f}%")
    ws7.column_dimensions['A'].width = 35

    # Save reports
    main_path   = os.path.join(output_dir, 'Automation_Test_Report.xlsx')
    passed_path = os.path.join(output_dir, 'Passed_Test_Cases.xlsx')
    failed_path = os.path.join(output_dir, 'Failed_Test_Cases.xlsx')
    summary_path= os.path.join(output_dir, 'Execution_Summary.xlsx')

    wb.save(main_path)

    # Passed-only workbook
    wb_p = Workbook()
    _write_test_sheet(wb_p.active, 'Passed Tests', passed, SIMPLE_COLS)
    wb_p.save(passed_path)

    # Failed-only workbook
    wb_f = Workbook()
    _write_test_sheet(wb_f.active, 'Failed Tests', failed, SIMPLE_COLS)
    wb_f.save(failed_path)

    # Summary workbook
    wb_s = Workbook()
    ws_s = wb_s.active
    ws_s.title = 'Execution Metrics'
    _cell_style(ws_s, 1, 1, 'Execution Metrics Summary', font=TITLE_FONT)
    _cell_style(ws_s, 2, 1, 'Metric', fill=HEADER_FILL, font=HEADER_FONT)
    _cell_style(ws_s, 2, 2, 'Value', fill=HEADER_FILL, font=HEADER_FONT)
    for i, (k, v) in enumerate(metric_data, 3):
        _cell_style(ws_s, i, 1, k)
        _cell_style(ws_s, i, 2, str(v))
    ws_s.column_dimensions['A'].width = 30
    ws_s.column_dimensions['B'].width = 30
    wb_s.save(summary_path)

    print(f"✅ Excel reports generated in: {output_dir}")
    return main_path


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', required=True, help='Path to execution-results.json')
    parser.add_argument('--output-dir', required=True, help='Output directory for Excel reports')
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"❌ Input file not found: {args.input}")
        sys.exit(1)

    with open(args.input, 'r') as f:
        results = json.load(f)

    generate_excel_report(results, args.output_dir)
